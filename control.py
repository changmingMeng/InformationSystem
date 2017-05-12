#! /usr/bin/python
# -*- coding: utf-8 -*-

import os.path
import csv
from openpyxl import Workbook
import datetime

from models import *

class control(object):

    @staticmethod
    def getListFromCSV(filepath):
        lst = []

        csvReader = csv.reader(file(filepath, 'rb'))
        for row in csvReader:
            for col in row:
                lst.append(col.decode("GBK").encode("utf-8"))

        return lst
#fn.SUM(CellBusi3G.erl), fn.SUM(CellBusi3G.updata)
    @classmethod
    def getZoneInfo(cls, begin_date, end_date, filepath):
        namelst = cls.getListFromCSV(filepath)
        sql = CellBusi3G.select(fn.SUM(CellBusi3G.erl), fn.SUM(CellBusi3G.updata), fn.SUM(CellBusi3G.downdata), fn.SUM(CellBusi3G.alldata))\
            .where((CellBusi3G.name << namelst)&(CellBusi3G.date.between(begin_date,end_date)))
        #sql = CellBusi3G.select().where(CellBusi3G.name << namelst).limit(50)
        return sql



    @staticmethod
    def getNobusiCell_prototype():
        """查询零业务小区的函数原型"""
        sql_get_ordered_table = "create temp table tem_cell_busi_2g as " \
                                "select name, date, erl, alldata " \
                                "from cell_busi_2g " \
                                "order by name, date;"
        execute_sql(sql_get_ordered_table)

        sql_get_lagged_table = "create temp table tem_lagged_cell_busi_2g as " \
                               "select *, lag(name,1) over(order by name)name_1, lag(name,2) over(order by name)name_2, lag(name,3) over(order by name)name_3, lag(name,4) over(order by name)name_4, " \
                               "lag(erl,1) over(order by name)erl_1 , lag(erl,2) over(order by name)erl_2, lag(erl,3) over(order by name)erl_3, lag(erl,4) over(order by name)erl_4 " \
                               "from tem_cell_busi_2g " \
                               "order by name, date;"
        execute_sql(sql_get_lagged_table)

        sql_get_zero_busi_cell = "select * " \
                                 "from tem_lagged_cell_busi_2g " \
                                 "where name=name_1 and name=name_2 and name=name_3 and name=name_4 and erl<0.01 and erl_1<0.01 and erl_2<0.01 and erl_3<0.01 and erl_4<0.01 " \
                                 "order by name, date;"
        return execute_sql(sql_get_zero_busi_cell)

    @staticmethod
    def getNobusiCell(days, threshold, nettype, busitype):
        """查询零业务小区，每次都生成临时表，效率较低"""
        lst_type = ['2g', '3g', '4g']
        lst_busitype = ['erl', 'data']
        if nettype not in lst_type:
            raise("wrong net type")
        if busitype not in lst_busitype:
            raise("wrong busi type")

        sql_get_ordered_table = "create temp table tem_cell_busi_"+nettype+" as " \
                                "select name, date, erl, alldata " \
                                "from cell_busi_"+nettype+" " \
                                "order by name, date;"
        print sql_get_ordered_table
        execute_sql(sql_get_ordered_table)

        def get_lagged_table(days, nettype):
            str_sql = "create temp table tem_lagged_cell_busi_"+nettype+" as select *"
            for i in range(1,days):
                str_sql += ", lag(name,"+str(i)+") over(order by name)name_"+str(i)
            for i in range(1,days):
                str_sql += ", lag(erl,"+str(i)+") over(order by name)erl_"+str(i)
            for i in range(1,days):
                str_sql += ", lag(alldata,"+str(i)+") over(order by name)alldata_"+str(i)
            str_sql += " from tem_cell_busi_"+nettype+" order by name, date;"
            return str_sql
        sql_get_lagged_table = get_lagged_table(days, nettype)
        print sql_get_lagged_table
        try:
            execute_sql(sql_get_lagged_table)
        except:
            execute_sql("drop table tem_cell_busi_" + nettype)
            raise("SQL error!")

        def get_zero_erl_cell(days, threshold, nettype):
            str_sql = "select name,min(date) from tem_lagged_cell_busi_"+nettype+" where"
            for i in range(1,days):
                str_sql += " name=name_"+str(i)+" and "
            if threshold == 0:
                str_sql += "erl=0 and "
                for i in range(1,days):
                    str_sql += "erl_"+str(i)+"=0 and "
            else:
                str_sql += "erl<"+str(threshold)+" and "
                for i in range(1,days):
                    str_sql += "erl_"+str(i)+"<"+str(threshold)+" and "
            str_sql += "true group by name order by name;"
            return str_sql

        def get_zero_data_cell(days, threshold, nettype):
            str_sql = "select name,min(date) from tem_lagged_cell_busi_" + nettype + " where"
            for i in range(1, days):
                str_sql += " name=name_" + str(i) + " and "
            if threshold == 0:
                str_sql += "alldata=0 and "
                for i in range(1, days):
                    str_sql += "alldata_" + str(i) + "=0 and "
            else:
                str_sql += "alldata<" + str(threshold) + " and "
                for i in range(1, days):
                    str_sql += "alldata_" + str(i) + "<" + str(threshold) + " and "
            str_sql += "true group by name order by name;"
            return str_sql
        if busitype == 'erl':
            sql_get_zero_busi_cell = get_zero_erl_cell(days, threshold, nettype)
        elif busitype == 'data':
            sql_get_zero_busi_cell = get_zero_data_cell(days, threshold, nettype)
        print sql_get_zero_busi_cell
        try:
            result = execute_sql(sql_get_zero_busi_cell).fetchall()
        except:
            execute_sql("drop table tem_cell_busi_" + nettype)
            execute_sql("drop table tem_lagged_cell_busi_" + nettype)
            raise ("SQL error!")
        #删除临时表以便下次查询
        execute_sql("drop table tem_cell_busi_" + nettype)
        execute_sql("drop table tem_lagged_cell_busi_" + nettype)
        return result

    @staticmethod
    def get_lagged_table(days, nettype):
        """利用postgreSQL的lag()函数，生成一张用户查询零业务小区的前置表"""
        str_sql = "create temp table tem_lagged_cell_busi_" + nettype + " as select *"
        for i in range(1, days):
            str_sql += ", lag(name," + str(i) + ") over(order by name)name_" + str(i)
        for i in range(1, days):
            str_sql += ", lag(erl," + str(i) + ") over(order by name)erl_" + str(i)
        for i in range(1, days):
            str_sql += ", lag(alldata," + str(i) + ") over(order by name)alldata_" + str(i)
        str_sql += " from tem_cell_busi_" + nettype + " order by name, date;"
        return str_sql

    @staticmethod
    def get_zero_erl_cell(days, threshold, nettype):
        str_sql = "select name,min(date) from tem_lagged_cell_busi_" + nettype + " where"
        for i in range(1, days):
            str_sql += " name=name_" + str(i) + " and "
        if threshold == 0:
            str_sql += "erl=0 and "
            for i in range(1, days):
                str_sql += "erl_" + str(i) + "=0 and "
        else:
            str_sql += "erl<" + str(threshold) + " and "
            for i in range(1, days):
                str_sql += "erl_" + str(i) + "<" + str(threshold) + " and "
        str_sql += "true group by name order by name;"
        return str_sql

    @staticmethod
    def get_zero_data_cell(days, threshold, nettype):
        str_sql = "select name,min(date) from tem_lagged_cell_busi_" + nettype + " where"
        for i in range(1, days):
            str_sql += " name=name_" + str(i) + " and "
        if threshold == 0:
            str_sql += "alldata=0 and "
            for i in range(1, days):
                str_sql += "alldata_" + str(i) + "=0 and "
        else:
            str_sql += "alldata<" + str(threshold) + " and "
            for i in range(1, days):
                str_sql += "alldata_" + str(i) + "<" + str(threshold) + " and "
        str_sql += "true group by name order by name;"
        return str_sql

    @staticmethod
    def get_zero_data_and_erl_cell(days, threshold, nettype):
        str_sql = "select name,min(date) from tem_lagged_cell_busi_" + nettype + " where"
        for i in range(1, days):
            str_sql += " name=name_" + str(i) + " and "
        if threshold == 0:
            str_sql += "alldata=0 and "
            for i in range(1, days):
                str_sql += "alldata_" + str(i) + "=0 and "

            str_sql += "and erl=0 and "
            for i in range(1, days):
                str_sql += "erl_" + str(i) + "=0 and "
        else:
            str_sql += "alldata<" + str(threshold) + " and "
            for i in range(1, days):
                str_sql += "alldata_" + str(i) + "<" + str(threshold) + " and "

            str_sql += "and erl<" + str(threshold) + " and "
            for i in range(1, days):
                str_sql += "erl_" + str(i) + "<" + str(threshold) + " and "
        str_sql += "true group by name order by name;"
        return str_sql

    @staticmethod
    def prepare_no_busi_cell(days=7):
        """生成查询零业务小区的临时前置表，服务器每天定时生成最近7天的，并删除前一天生成的临时表"""
        lst_type = ['2g', '3g', '4g']

        today = datetime.datetime.now().date()
        begindate = today - datetime.timedelta(days=2+days)
        enddate = today - datetime.timedelta(days=2)
        for nettype in lst_type:
            # 删除前一天生成的临时表
            try:
                execute_sql("drop table tem_cell_busi_" + nettype)
                execute_sql("drop table tem_lagged_cell_busi_" + nettype)
            except:
                pass

            sql_get_ordered_table = "create temp table tem_cell_busi_" \
                                    + nettype \
                                    + " as select name, date, erl, alldata "\
                                      "from cell_busi_" + nettype + " " \
                                      "where date between '" + str(begindate) + "' and '" + str(enddate) + "' "\
                                      "order by name, date;"
            print sql_get_ordered_table
            execute_sql(sql_get_ordered_table)

            sql_get_lagged_table = control.get_lagged_table(days, nettype)
            print sql_get_lagged_table
            try:
                execute_sql(sql_get_lagged_table)
            except:
                execute_sql("drop table tem_cell_busi_" + nettype)
                raise ("SQL error!")

    @staticmethod
    def quick_zero_busi_cell(busitype, days, threshold, nettype):

        if busitype == 'erl':
            sql_get_zero_busi_cell = control.get_zero_erl_cell(days, threshold, nettype)
        elif busitype == 'data':
            sql_get_zero_busi_cell = control.get_zero_data_cell(days, threshold, nettype)
        elif busitype == 'both':
            sql_get_zero_busi_cell = control.get_zero_data_and_erl_cell(days, threshold, nettype)
        else:
            raise("busi type error")
        print sql_get_zero_busi_cell
        try:
            result = execute_sql(sql_get_zero_busi_cell).fetchall()
        except:
            raise ("zero busi cell SQL error")

        return result


    @staticmethod
    def write2Excel(lst_result, filepath):
        wb = Workbook()
        ws = wb.active

        r = 0
        for t in lst_result:
            r += 1
            ws.cell(row=r, column=1, value=t[0])
            ws.cell(row=r, column=2, value=t[1])
        print r

        wb.save(filepath)

def test(filepath):
    print control.getListFromCSV(filepath)
    ctl = control()
    sql = ctl.getZoneInfo("2017-03-01", "2017-03-02", filepath)
    if sql is not None:
        for t in sql:
            print t.erl, t.updata, t.downdata, t.alldata
    else:
        print "sql is None"

if __name__ == "__main__":
    #test(r"E:\developtools\PyCharm 2016.3\projects\InformationSystem\files\test.csv")
    lst_result = control.getNobusiCell(3, 0.01, '2g', 'data')
    print datetime.datetime.now().date()
